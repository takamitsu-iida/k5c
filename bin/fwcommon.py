#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
ファイアウォール関係の共通処理
"""

import logging
import sys

try:
  from tabulate import tabulate
except ImportError as e:
  logging.exception("tabulateモジュールのインポートに失敗しました: %s", e)
  sys.exit(1)

try:
  from openpyxl import load_workbook
  # from openpyxl.utils import column_index_from_string
except ImportError as e:
  logging.exception("openpyxlモジュールのインポートに失敗しました: %s", e)
  sys.exit(1)


#
# create, update, showの結果表示
#
def print_fw_rule(data=None):
  """結果を表示します"""

  # 作成したルールの情報はデータオブジェクトの中の'firewall_rule'キーにオブジェクトとして入っている
  #"data": {
  #  "firewall_rule": {
  #    "destination_ip_address": null,
  #    "action": "allow",
  #    "ip_version": 4,
  #    "firewall_policy_id": null,
  #    "position": null,
  #    "source_ip_address": "133.162.192.0/24",
  #    "id": "de2bb711-d495-4ae5-9d05-672575d1549a",
  #    "shared": false,
  #    "availability_zone": "jp-east-1a",
  #    "destination_port": null,
  #    "enabled": true,
  #    "protocol": "tcp",
  #    "description": "",
  #    "name": "iida-az1-p01-mgmt01-any-tcp",
  #    "tenant_id": "a5001a8b9c4a4712985c11377bd6d4fe",
  #    "source_port": null
  #  }
  #}
  rule = data.get('firewall_rule', {})

  # 表示用に配列にする
  disp_keys = [
    'id', 'name', 'enabled', 'action', 'protocol',
    'source_ip_address', 'source_port', 'destination_ip_address', 'destination_port',
    'description', 'availability_zone', 'tenant_id'
  ]

  rules = []
  for key in disp_keys:
    rules.append([key, rule.get(key, '')])

  # ファイアウォールルール情報を表示
  print("/v2.0/fw/firewall_rules")
  print(tabulate(rules, tablefmt='rst'))


#
# create, update, showの結果表示
#
def print_fw_policy(data=None):
  """結果を表示します"""

  # 作成したポリシーの情報はデータオブジェクトの中の'firewall_policy'キーにオブジェクトとして入っている
  #"data": {
  #  "firewall_policy": {
  #    "id": "84417ab6-53ea-4595-9d92-7a9d9a552e12",
  #    "tenant_id": "a5001a8b9c4a4712985c11377bd6d4fe",
  #    "firewall_rules": [
  #      "867e35c5-2875-4c51-af31-4cf7932f17f6",
  #      "9597ea56-e39d-4160-bdca-ebf2aca23aab",
  #      "589d96ad-79e9-4a84-b923-10145469643c",
  #      "c44321b7-6b04-4ec4-8e62-dc080794f59b",
  #      "57fbe4aa-6edf-4123-8b6c-c8233cfb3c70",
  #      "6eb8b10f-0756-460a-8b6a-8dd3db77173d",
  #      "58dfdc2f-bf23-481b-9f3a-4b96df6232a2",
  #      "75877f59-7a26-4a59-a343-9e2955dfb49e",
  #      "8cf91195-d611-489d-b322-e28cab2ba705",
  #      "acfada3d-0527-43e7-ba4d-6403ca8654fe",
  #      "3750f08f-5567-4ad7-870f-dd830cc898b0",
  #      "bc8f66e6-c09c-448f-869c-96f2c0843e81"
  #    ],  #    "description": "",
  #    "availability_zone": "jp-east-1a",
  #    "shared": false,
  #    "name": "iida",
  #    "audited": false
  #  }
  #},
  fp = data.get('firewall_policy', {})

  # 表示用に配列にする
  disp_keys = ['id', 'name', 'availability_zone', 'tenant_id']
  fps = []
  for key in disp_keys:
    fps.append([key, fp.get(key, '')])

  # ファイアウォールポリシー情報を表示
  print("/v2.0/fw/firewall_policies")
  print(tabulate(fps, tablefmt='rst'))

  # ルール一覧を表示
  rules = []
  for item in fp.get('firewall_rules', []):
    rules.append([item])
  print(tabulate(rules, tablefmt='rst'))


def get_rules_to_create(filename=""):
  """エクセルファイルからIDの割り当てられていないルールオブジェクトを配列にして返します"""

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return None
  name_column_letter = cell.column  # C
  # name_column_index = column_index_from_string(name_column_letter)  # C -> 2
  name_row_index = cell.row  # 7

  # 同じ行から 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='id')
  if not cell:
    return None
  id_column_letter = cell.column  # D

  # 結果配列
  result_list = []

  # 名前が記載された列を上から順に確認する
  for cell in ws[name_column_letter]:
    if cell.row <= name_row_index:
      continue
    if cell.row > 1024:
      break
    if not cell.value:
      continue

    # IDが既に記載されているものは飛ばす
    if ws[id_column_letter + str(cell.row)].value:
      continue

    # ルールが記載された行を取り出す
    rule_row = ws[cell.row]

    # 作成用に許可されたキーの一覧
    allowed_key_list = [
      'name', 'enabled', 'action', 'ip_version', 'protocol', 'source_ip_address', 'source_port',
      'destination_ip_address', 'destination_port', 'description', 'availability_zone'
    ]

    # nameと同じ行からキーを拾いながらオブジェクトを作成する
    # POSTで新規作成するときはpositionやidキーを含めてはならないので取り除く
    rule_object = {}
    for cell in rule_row:
      key = ws[cell.column + str(name_row_index)].value
      if not key in allowed_key_list:
        continue
      value = cell.value
      # quick hack, description key must be string
      if key == 'description' and not value:
        value = ''
      if str(value).upper() == 'NULL':
        value = None
      rule_object[key] = value

    result_list.append(rule_object)

  return result_list


def get_rule_to_update(filename="", rule_id=""):
  """エクセルファイルから指定されたIDのルールオブジェクトを作成して返却します"""

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return None
  # name_column_letter = cell.column  # C
  # name_column_index = column_index_from_string(name_column_letter)  # C -> 2
  name_row = cell.row  # 7

  # 同じ行から 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row], value='id')
  if not cell:
    return None
  id_column_letter = cell.column  # D

  # 'id' 列を上から舐めて、指定された名前のものがあるか探す
  cell = find_cell(col=ws[id_column_letter], value=rule_id)
  if not cell:
    return None
  rule_row_index = cell.row

  # ルールが記載された行を取り出す
  rule_row = ws[rule_row_index]

  # 返却値
  result = {}

  # update用に許可されたキーの一覧
  allowed_key_list = [
    'name', 'enabled', 'action', 'protocol', 'source_ip_address', 'source_port',
    'destination_ip_address', 'destination_port'
  ]

  # nameと同じ行からキーを拾いながら戻り値のオブジェクトを作成する
  # POSTで新規作成するときはpositionやidキーを含めてはならないので取り除く
  for cell in rule_row:
    key = ws[cell.column + str(name_row)].value
    if not key in allowed_key_list:
      continue
    value = cell.value
    # quick hack, description key must be string
    if key == 'description' and not value:
      value = ''
    if str(value).upper() == 'NULL':
      value = None
    result[key] = value

  return result


def get_rule_id_list(filename=""):
  """エクセルファイルからルールIDの一覧を読んで返却します"""

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return None
  # name_column_letter = cell.column  # C
  # name_column_index = column_index_from_string(name_column_letter)  # C -> 2
  name_row_index = cell.row  # 7

  # 同じ行から 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='id')
  if not cell:
    return None
  id_column_letter = cell.column  # D

  # 同じ行から 'firewall_policy_id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='firewall_policy_id')
  if not cell:
    return None
  policy_column_letter = cell.column  # O

  # 返却値
  result = []

  # ルールIDが記載された列からid値を取り出す
  for cell in ws[id_column_letter]:
    if cell.row <= name_row_index:
      continue
    if cell.row > 1024:
      break
    if not cell.value:
      continue
    if ws[policy_column_letter + str(cell.row)].value:
      continue
    result.append(cell.value)

  return result


def save_policy(filename="", rule_id_list=None, policy_id="", policy_name=""):
  """ポリシーIDと名前をファイルに書き込みます"""
  if not rule_id_list or not policy_id:
    return

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return
  name_row_index = cell.row  # 7

  # 同じ行の 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='id')
  if not cell:
    return
  id_column_letter = cell.column  # O

  # 同じ行の 'firewall_policy_id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='firewall_policy_id')
  if not cell:
    return
  policy_id_column_letter = cell.column  # O

  # 同じ行の 'firewall_policy_name' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='firewall_policy_name')
  if not cell:
    return
  policy_name_column_letter = cell.column  # P

  # 'id' 列を上から舐めながらfirewall_policy_id列に値をセットする
  for cell in ws[id_column_letter]:
    if cell.row <= name_row_index:
      continue
    # 1024行に達したらそれ以上は無駄なので抜ける
    if cell.row > 1024:
      break
    rule_id = ws[id_column_letter + str(cell.row)].value
    if rule_id in rule_id_list:
      ws[policy_id_column_letter + str(cell.row)].value = policy_id
      ws[policy_name_column_letter + str(cell.row)].value = policy_name

  # ファイルを保存
  try:
    wb.save(filename)
  except PermissionError as e:
    logging.error(e)


def save_rule(filename="", name="", rule_id=None):
  """ルールIDをファイルに書き込みます"""
  if not rule_id:
    return

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return
  name_column_letter = cell.column  # C
  name_row = cell.row  # 7

  # 同じ行の 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row], value='id')
  if not cell:
    return
  id_column_letter = cell.column  # D

  # 'name' 列を上から舐めて、指定された名前のものがあるか探す
  cell = None
  for row in ws[name_column_letter]:
    # 一致したら抜ける
    if row.value == name:
      cell = row
      break
    # 1024行に達したらそれ以上は無駄なので抜ける
    if row.row > 1024:
      break
  if not cell:
    return None

  # 値をセットして
  ws[id_column_letter + str(cell.row)].value = rule_id

  # ファイルを保存
  try:
    wb.save(filename)
  except PermissionError as e:
    logging.error("Failed to save excel file.")


def find_cell(ws=None, row=None, col=None, value=None):
  """指定された値を探します"""
  # 行が指定されているならその行から探す
  if row:
    for cell in row:
      if cell.value == value:
        return cell
    return None

  if col:
    for cell in col:
      if cell.value == value:
        return cell
      # 1024行に達したらそれ以上は無駄なので抜ける
      if cell.row > 1024:
        break
    return None

  # ワークシートから探す
  # 1行~1024行、26列までを探索
  for row in ws.iter_rows(min_row=1, max_row=1024, max_col=26):
    for cell in row:
      if cell.value == value:
        return cell
  return None


def get_policy_rules_to_update(filename="", policy_id=""):
  """エクセルファイルから指定されたポリシーIDのルールID配列を作成して返却します"""

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.error(e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return None
  # name_column_letter = cell.column  # C
  # name_column_index = column_index_from_string(name_column_letter)  # C -> 2
  name_row_index = cell.row  # 7

  # 同じ行から 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='id')
  if not cell:
    return None
  id_column_letter = cell.column  # D

  # 同じ行から 'firewall_policy_id' という値のセルを探す
  cell = find_cell(row=ws[name_row_index], value='firewall_policy_id')
  if not cell:
    return None
  firewall_policy_id_column_letter = cell.column  # P

  # 返却値
  result = []

  # firewall_policy_id列からid値を取り出す
  for cell in ws[firewall_policy_id_column_letter]:
    if cell.row <= name_row_index:
      continue
    if cell.row > 1024:
      break
    if not cell.value:
      continue
    if cell.value == policy_id:
      rule_id = ws[id_column_letter + str(cell.row)].value
      if rule_id:
        result.append(rule_id)

  return result
