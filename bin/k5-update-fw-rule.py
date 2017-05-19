#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
PUT /v2.0/fw/firewall_rules/{firewall_rule_id}
Update firewall rule
ファイアーウォールのルールを更新する

NOTE:
　・ルールはエクセルで作成します
　・ルールファイルのデフォルトのファイル名は$app_home/conf/fw-rules.xlsxです
　・エクセル内の同じidのルールを変更します
"""

"""
実行例

bash-4.4$ ./bin/k5-update-fw-rule.py --rule-id 2b9a71f6-6885-460e-a1a4-32e84c126a0e
PUT /v2.0/fw/firewall_rules
======================  ====================================
id                      2b9a71f6-6885-460e-a1a4-32e84c126a0e
name                    iida-az1-p01-net02-any-tcp
enabled                 True
action                  allow
protocol                tcp
source_ip_address       10.1.2.0/24
source_port
destination_ip_address
destination_port
description
availability_zone       jp-east-1a
tenant_id               a5001a8b9c4a4712985c11377bd6d4fe
======================  ====================================
"""

import json
import logging
import os
import sys

def here(path=''):
  """相対パスを絶対パスに変換して返却します"""
  return os.path.abspath(os.path.join(os.path.dirname(__file__), path))

# libフォルダにおいたpythonスクリプトを読みこませるための処理
if not here("../lib") in sys.path:
  sys.path.append(here("../lib"))

if not here("../lib/site-packages") in sys.path:
  sys.path.append(here("../lib/site-packages"))

try:
  from k5c import k5c
except ImportError as e:
  logging.exception("k5cモジュールのインポートに失敗しました: %s", e)
  exit(1)

try:
  from tabulate import tabulate
except ImportError as e:
  logging.exception("tabulateモジュールのインポートに失敗しました: %s", e)
  exit(1)

try:
  from openpyxl import load_workbook
  # from openpyxl.utils import column_index_from_string
except ImportError as e:
  logging.exception("openpyxlモジュールのインポートに失敗しました: %s", e)
  exit(1)


#
# リクエストデータを作成
#
def make_request_data(filename="", rule_id=""):
  """リクエストデータを作成して返却します"""
  rule = read_rule_by_id(filename=filename, rule_id=rule_id)
  if not rule:
    return None

  # 作成するルールのオブジェクト
  request_data = {
    'firewall_rule': rule
  }

  return request_data


#
# APIにアクセス
#
def access_api(rule_id="", data=None):
  """REST APIにアクセスします"""

  # 接続先URL
  url = k5c.EP_NETWORK + "/v2.0/fw/firewall_rules/" + rule_id

  # Clientクラスをインスタンス化
  c = k5c.Client()

  # PUTメソッドで更新
  r = c.put(url=url, data=data)

  return r


#
# 結果を表示
#
def print_result(result=None, dump=False):
  """結果を表示します"""

  # 中身を確認
  if dump:
    print(json.dumps(result, indent=2))
    return

  # ステータスコードは'status_code'キーに格納
  status_code = result.get('status_code', -1)

  # ステータスコードが異常な場合
  if status_code < 0 or status_code >= 400:
    print(json.dumps(result, indent=2))
    return

  # データは'data'キーに格納
  data = result.get('data', None)
  if not data:
    logging.error("no data found")
    return

  # 作成したルールの情報はデータオブジェクトの中の'firewall_rule'キーにオブジェクトとして入っている
  #"data": {
  #  "firewall_rule": {
  #    "destination_ip_address": null,
  #    "action": "allow",
  #    "ip_version": 4,
  #    "firewall_policy_id": null,
  #    "position": null,
  #    "source_ip_address": "133.162.192.0/24",
  #    "id": "de2bb711-d495-4ae5-9d05-672575d1549a",
  #    "shared": false,
  #    "availability_zone": "jp-east-1a",
  #    "destination_port": null,
  #    "enabled": true,
  #    "protocol": "tcp",
  #    "description": "",
  #    "name": "iida-az1-p01-mgmt01-any-tcp",
  #    "tenant_id": "a5001a8b9c4a4712985c11377bd6d4fe",
  #    "source_port": null
  #  }
  #}
  rule = data.get('firewall_rule', {})

  # 表示用に配列にする
  rules = []
  rule_id = rule.get('id', '')
  rules.append(['id', rule_id])
  rules.append(['name', rule.get('name', '')])
  rules.append(['enabled', rule.get('enabled', '')])
  rules.append(['action', rule.get('action', '')])
  rules.append(['protocol', rule.get('protocol', '')])
  rules.append(['source_ip_address', rule.get('source_ip_address', '')])
  rules.append(['source_port', rule.get('source_port', '')])
  rules.append(['destination_ip_address', rule.get('destination_ip_address', '')])
  rules.append(['destination_port', rule.get('destination_port', '')])
  rules.append(['description', rule.get('description', '')])
  rules.append(['availability_zone', rule.get('availability_zone', '')])
  rules.append(['tenant_id', rule.get('tenant_id', '')])

  # ファイアウォールルール情報を表示
  print("PUT /v2.0/fw/firewall_rules")
  print(tabulate(rules, tablefmt='rst'))


def get_rule_id(result=None):
  """rule_idを探して返却します"""
  data = result.get('data', None)
  if not data:
    return None
  rule = data.get('firewall_rule', {})
  return rule.get('id', '')


def read_rule_by_id(filename="", rule_id=""):
  """ファイルからルールを読みます"""

  # Excelのブックファイルを読み出す
  try:
    wb = load_workbook(filename=filename, data_only=True)
  except FileNotFoundError as e:
    logging.exception("ファイルが見つかりません: %s", e)
    return None

  # アクティブなシートを取り出す
  ws = wb.active

  # 'name' という値のセルを探す
  cell = find_cell(ws=ws, value='name')
  if not cell:
    return None
  # name_column_letter = cell.column  # C
  # name_column_index = column_index_from_string(name_column_letter)  # C -> 2
  name_row = cell.row  # 7

  # 同じ行から 'id' という値のセルを探す
  cell = find_cell(row=ws[name_row], value='id')
  if not cell:
    return None
  id_column_letter = cell.column  # D

  # 'id' 列を上から舐めて、指定された名前のものがあるか探す
  cell = find_cell(col=ws[id_column_letter], value=rule_id)
  if not cell:
    return None
  rule_row_index = cell.row

  # ルールが記載された行を取り出す
  rule_row = ws[rule_row_index]

  # 返却値
  result = {}

  # nameと同じ行からキーを拾いながら戻り値のオブジェクトを作成する
  # POSTで新規作成するときはpositionやidキーを含めてはならないので取り除く
  for cell in rule_row:
    key = ws[cell.column + str(name_row)].value
    if not key_allowed(key):
      continue
    value = cell.value
    # quick hack, description key must be string
    if key == 'description' and not value:
      value = ''
    if str(value).upper() == 'NULL':
      value = None
    result[key] = value

  return result


def key_allowed(key=''):
  """許可されたキーであればTrue、その他はFalseを返します"""
  if not key:
    return False

  allowed_keys = [
    'name', 'enabled', 'action', 'protocol', 'source_ip_address', 'source_port',
    'destination_ip_address', 'destination_port'
  ]

  if key.lower() in allowed_keys:
    return True
  return False


def find_cell(ws=None, row=None, col=None, value=None):
  """指定された値を探します"""
  # 行が指定されているならその行から探す
  if row:
    for cell in row:
      if cell.value == value:
        return cell
    return None

  if col:
    for cell in col:
      if cell.value == value:
        return cell
      # 1024行に達したらそれ以上は無駄なので抜ける
      if cell.row > 1024:
        break
    return None

  # ワークシートから探す
  # 1行~1024行、26列までを探索
  for row in ws.iter_rows(min_row=1, max_row=1024, max_col=26):
    for cell in row:
      if cell.value == value:
        return cell
  return None


if __name__ == '__main__':

  import argparse

  def main():
    """メイン関数"""

    # アプリケーションのホームディレクトリ
    app_home = here("..")

    # 設定ファイルのパス
    # $app_home/conf/fw-rules.xlsx
    config_file = os.path.join(app_home, "conf", "fw-rules.xlsx")

    parser = argparse.ArgumentParser(description='Updates a firewall rule.')
    parser.add_argument('--rule-id', dest='rule_id', metavar='id', required=True, help='The rule id.')
    parser.add_argument('--filename', metavar='file', default=config_file, help='The rule file. default: '+config_file)
    parser.add_argument('--dump', action='store_true', default=False, help='Dump json result and exit.')
    args = parser.parse_args()
    rule_id = args.rule_id
    filename = args.filename
    dump = args.dump

    # ファイルからルールを読み取ってリクエストデータを作る
    data = make_request_data(filename=filename, rule_id=rule_id)
    # print(json.dumps(data, indent=2))

    if not data:
      logging.error('no rule found.')
      return

    # 実行
    result = access_api(rule_id=rule_id, data=data)

    # 得たデータを処理する
    print_result(result=result, dump=dump)


  # 実行
  main()
